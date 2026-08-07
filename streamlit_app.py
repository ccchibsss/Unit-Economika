Полный скрипт целиком, без сокращений. Интегрированы строгие формулы FBS Яндекс Маркета (Прибыль = P × (1 − Com% − Pay% − Tax% − Ad%) − C − Pack − Log − Proc − Other − Ret% × RetCost), лишние столбцы и расчёты (объёмный вес, магистраль, ABC/XYZ, спецтарифы) удалены. Расчёты синхронизированы между приложением и Excel-экспортом.

```python
#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
============================================================================
🚀 YANDEX MARKET FBS UNIT ECONOMICS v24.0 — CLEAN FORMULAS EDITION
============================================================================
Интегрированы точные формулы юнит-экономики FBS:

    Комиссия              = P × Com%
    Эквайринг             = P × Pay%
    Реклама (ДРР)         = P × Ad%
    Налог (УСН доходы)    = P × Tax%
    Риск возвратов        = Ret% × RetCost

    Чистая прибыль = P × (1 − Com% − Pay% − Tax% − Ad%)
                     − C − Pack − Log − Proc − Other − Ret% × RetCost

    Маржинальность = Прибыль / P × 100%
    Наценка        = Прибыль / C × 100%

    Pmin (цена в ноль) = (C + Pack + Log + Proc + Other + Ret% × RetCost)
                         / (1 − Com% − Pay% − Tax% − Ad%)

    Цена для маржи M%  = (C + Pack + Log + Proc + Other + Ret% × RetCost)
                         / (1 − Com% − Pay% − Tax% − Ad% − M%)

Обозначения переменных:
    P       — цена продажи (selling_price)
    C       — себестоимость (cogs)
    Pack    — стоимость упаковки (packaging_cost)
    Com%    — комиссия Маркета по категории (commission_rate)
    Log     — логистика FBS / доставка до покупателя (delivery_cost)
    Proc    — обработка отправления (processing_cost)
    Pay%    — эквайринг / приём платежей (acquiring_rate)
    Ad%     — реклама, ДРР, бусты (advertising_rate)
    Tax%    — налоговая ставка (tax_rate)
    Ret%    — доля возвратов / невыкупов (return_rate)
    RetCost — средняя стоимость одного возврата (return_cost_amount)
    Other   — прочие расходы: маркировка, стикеры, доставка до ПВЗ (other_costs)
============================================================================
"""
import streamlit as st
import pandas as pd
import numpy as np
import io
import hashlib
import logging
import warnings
from datetime import datetime
from decimal import Decimal, ROUND_HALF_UP
from typing import Dict, List, Optional, Tuple, Any, TypedDict, Final
from dataclasses import dataclass
from collections import OrderedDict
import time
from contextlib import contextmanager

import requests
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from openpyxl.chart import Reference, PieChart
    from openpyxl.chart.label import DataLabelList
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

warnings.filterwarnings('ignore')
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger('YandexMarketUE')

# ============================================================================
# КОНСТАНТЫ
# ============================================================================
APP_VERSION: Final[str] = "24.0.0"
APP_NAME: Final[str] = "Yandex Market FBS Unit Economics"
CACHE_TTL: Final[int] = 3600
LRU_CACHE_SIZE: Final[int] = 128
MAX_RETRIES: Final[int] = 3
REQUEST_TIMEOUT: Final[int] = 15

# ============================================================================
# ТИПИЗИРОВАННЫЕ КОНФИГУРАЦИИ
# ============================================================================
class TariffDict(TypedDict, total=False):
    category: str
    commission_rate: float
    delivery_cost: float
    processing_cost: float
    source: str

# ============================================================================
# УТИЛИТЫ (БЛОК 0)
# ============================================================================
class NumericUtils:
    """Утилиты для точных денежных расчётов."""

    @staticmethod
    @np.vectorize
    def money_round(value: float) -> float:
        """Округление до 2 знаков с защитой от NaN/Inf."""
        if not np.isfinite(value):
            return 0.0
        return float(Decimal(str(value)).quantize(
            Decimal("0.00"), rounding=ROUND_HALF_UP
        ))

    @staticmethod
    @np.vectorize
    def percent_round(value: float) -> float:
        """Округление процентов до 2 знаков."""
        if not np.isfinite(value):
            return 0.0
        return float(Decimal(str(value)).quantize(
            Decimal("0.01"), rounding=ROUND_HALF_UP
        ))

    @staticmethod
    def safe_divide(
        numerator: np.ndarray,
        denominator: np.ndarray,
        default: float = 0.0
    ) -> np.ndarray:
        """Безопасное деление массивов."""
        numerator = np.asarray(numerator, dtype=np.float64)
        denominator = np.asarray(denominator, dtype=np.float64)
        mask = (np.abs(denominator) < 1e-10) | ~np.isfinite(denominator) | ~np.isfinite(numerator)
        result = np.divide(numerator, denominator, where=~mask)
        result[mask] = default
        return result


class DtypeOptimizer:
    """Оптимизация типов данных для экономии памяти."""

    FLOAT_COLS = {
        'selling_price', 'cogs', 'packaging_cost', 'delivery_cost',
        'processing_cost', 'other_costs', 'commission', 'acquiring_cost',
        'advertising_cost', 'tax_cost', 'expected_return_cost',
        'total_expenses', 'gross_profit', 'rec_price_min',
        'rec_price_20', 'rec_price_30'
    }

    @classmethod
    def optimize(cls, df: pd.DataFrame) -> pd.DataFrame:
        """Оптимизация типов данных DataFrame."""
        df = df.copy()

        # Строковые колонки → category dtype
        for col in ['artikul', 'category', 'profitability_status']:
            if col in df.columns:
                df[col] = df[col].astype('category')

        # Float колонки → float32
        for col in cls.FLOAT_COLS:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0.0)
                df[col] = df[col].astype(np.float32)

        return df


class StringUtils:
    """Утилиты для работы со строками."""

    @staticmethod
    def fix_double_utf8(text: str) -> str:
        """Исправление двойного кодирования UTF-8."""
        if not isinstance(text, str) or not text:
            return text

        for source_enc, target_enc in [('cp1251', 'utf-8'), ('latin1', 'utf-8')]:
            try:
                fixed = text.encode(source_enc).decode(target_enc)
                if fixed and 'Ð' not in fixed[:2]:
                    return fixed
            except (UnicodeEncodeError, UnicodeDecodeError):
                continue
        return text

    @staticmethod
    def make_hash(obj: Any) -> str:
        """Детерминированный хеш для кэширования."""
        try:
            if isinstance(obj, pd.DataFrame):
                return hashlib.sha256(
                    pd.util.hash_pandas_object(obj, index=True).values.tobytes()
                ).hexdigest()[:16]
            return hashlib.sha256(str(obj).encode()).hexdigest()[:16]
        except Exception:
            return hashlib.sha256(b"hash_fallback").hexdigest()[:16]


class LRUCache:
    """LRU-кэш с TTL для тарифов."""

    def __init__(self, max_size: int = LRU_CACHE_SIZE, ttl: int = CACHE_TTL):
        self.max_size = max_size
        self.ttl = ttl
        self._cache: OrderedDict[str, Tuple[Any, float]] = OrderedDict()

    def get(self, key: str) -> Optional[Any]:
        """Получение значения из кэша."""
        if key not in self._cache:
            return None

        value, timestamp = self._cache[key]
        if time.time() - timestamp > self.ttl:
            del self._cache[key]
            return None

        self._cache.move_to_end(key)
        return value

    def set(self, key: str, value: Any) -> None:
        """Сохранение значения в кэш."""
        if key in self._cache:
            self._cache.move_to_end(key)

        self._cache[key] = (value, time.time())

        if len(self._cache) > self.max_size:
            self._cache.popitem(last=False)

    def clear(self) -> None:
        """Очистка кэша."""
        self._cache.clear()

# ============================================================================
# МОДЕЛИ (БЛОК 1)
# ============================================================================
@dataclass
class Tariff:
    """
    Модель тарифа категории по блокам тарифов Яндекс Маркета FBS:
    - commission_rate  → Комиссия за продажу (Com%), зависит от категории
    - delivery_cost    → Логистика / доставка до покупателя (Log), ₽
    - processing_cost  → Обработка отправления (Proc), ₽
    Эквайринг, реклама, налог и возвраты задаются глобально в настройках.
    """
    category: str
    commission_rate: float = 0.15
    delivery_cost: float = 120.0
    processing_cost: float = 45.0
    source: str = "Базовый фоллбэк"

    def __post_init__(self):
        """Валидация значений после инициализации."""
        self.category = str(self.category).lower().strip()
        self.commission_rate = max(0.0, float(self.commission_rate))
        self.delivery_cost = max(0.0, float(self.delivery_cost))
        self.processing_cost = max(0.0, float(self.processing_cost))

    def to_dict(self) -> TariffDict:
        """Сериализация в словарь."""
        return {
            'category': self.category,
            'commission_rate': self.commission_rate,
            'delivery_cost': self.delivery_cost,
            'processing_cost': self.processing_cost,
            'source': self.source
        }

    @classmethod
    def default(cls, category: str = "default") -> 'Tariff':
        """Создание тарифа по умолчанию."""
        return cls(category=category)

# ============================================================================
# API КЛИЕНТ (БЛОК 2)
# ============================================================================
class RateLimiter:
    """Простой rate limiter для API запросов."""

    def __init__(self, max_calls: int = 10, period: float = 1.0):
        self.max_calls = max_calls
        self.period = period
        self.calls: List[float] = []

    def wait_if_needed(self) -> None:
        """Ожидание при превышении лимита."""
        now = time.time()
        self.calls = [t for t in self.calls if now - t < self.period]

        if len(self.calls) >= self.max_calls:
            sleep_time = self.period - (now - self.calls[0])
            if sleep_time > 0:
                time.sleep(sleep_time)

        self.calls.append(time.time())


class APIClient:
    """HTTP-клиент с retry, backoff и rate limiting."""

    def __init__(
        self,
        base_url: str,
        api_key: str,
        max_retries: int = MAX_RETRIES,
        timeout: int = REQUEST_TIMEOUT
    ):
        self.base_url = base_url.rstrip('/')
        self.api_key = api_key
        self.timeout = timeout
        self.rate_limiter = RateLimiter()

        self.session = requests.Session()
        retry_strategy = Retry(
            total=max_retries,
            backoff_factor=1,
            status_forcelist=[429, 500, 502, 503, 504],
            allowed_methods=["HEAD", "GET", "POST", "PUT", "DELETE", "OPTIONS", "TRACE"]
        )
        adapter = HTTPAdapter(max_retries=retry_strategy)
        self.session.mount("https://", adapter)
        self.session.mount("http://", adapter)

    @contextmanager
    def _request_context(self):
        """Контекстный менеджер для запросов."""
        self.rate_limiter.wait_if_needed()
        try:
            yield
        finally:
            pass

    def _request(
        self,
        method: str,
        endpoint: str,
        **kwargs
    ) -> Dict[str, Any]:
        """Выполнение HTTP-запроса с retry логикой."""
        if not self.api_key:
            return {}

        url = f"{self.base_url}/{endpoint.lstrip('/')}"

        with self._request_context():
            try:
                resp = self.session.request(
                    method,
                    url,
                    timeout=self.timeout,
                    **kwargs
                )
                resp.raise_for_status()
                return resp.json()
            except requests.exceptions.Timeout:
                logger.warning(f"Таймаут {self.timeout}s: {url}")
            except requests.exceptions.HTTPError as e:
                logger.warning(
                    f"HTTP {e.response.status_code}: {e.response.text[:200]}"
                )
            except Exception as e:
                logger.warning(f"Ошибка запроса {url}: {e}")

        return {}


class YandexMarketAPI(APIClient):
    """API-клиент Яндекс Маркета."""

    def __init__(self, api_key: str, business_id: Optional[str] = None):
        super().__init__("https://api.partner.market.yandex.ru", api_key)
        self.business_id = business_id
        self.headers = {
            "Authorization": f"Bearer {api_key}",
            "Content-Type": "application/json",
            "Accept": "application/json"
        }
        if business_id:
            self.headers["X-Business-Id"] = business_id

    def get_campaigns(self) -> List[Dict[str, Any]]:
        """Получение списка кампаний."""
        data = self._request("GET", "/campaigns", headers=self.headers)
        return data.get("campaigns", [])

    def calculate_tariffs(
        self,
        offers: List[Dict[str, Any]],
        campaign_id: Optional[int] = None,
        selling_program: str = "FBS"
    ) -> List[Dict[str, Any]]:
        """Расчёт тарифов для списка офферов."""
        payload = {
            "parameters": {
                "sellingProgram": selling_program,
                "frequency": "WEEKLY",
                "paymentDelayWeeks": 4,
                "currency": "RUR"
            },
            "offers": offers
        }
        if campaign_id:
            payload["parameters"]["campaignId"] = campaign_id
            del payload["parameters"]["sellingProgram"]

        data = self._request(
            "POST",
            "/tariffs/calculate",
            headers=self.headers,
            json=payload
        )
        return data.get("result", {}).get("offers", [])

# ============================================================================
# МЕНЕДЖЕР ТАРИФОВ (БЛОК 3)
# ============================================================================
class HybridTariffManager:
    """Управление тарифами категорий с LRU-кэшем."""

    def __init__(self):
        if 'tariffs' not in st.session_state:
            st.session_state.tariffs = {}

        self._cache = LRUCache()

    @property
    def tariffs(self) -> Dict[str, Tariff]:
        """Загруженные тарифы."""
        return st.session_state.tariffs

    def load_tariffs_from_file(self, df: pd.DataFrame) -> int:
        """
        Загрузка тарифов из DataFrame.
        Ожидаемые колонки: category, commission_rate, delivery_cost, processing_cost
        """
        required_cols = {'category', 'commission_rate'}
        if not required_cols.issubset(df.columns):
            raise ValueError(
                f"Файл тарифов должен содержать: {required_cols}"
            )

        loaded = 0
        for _, row in df.iterrows():
            cat = str(row.get('category', '')).lower().strip()
            if not cat or cat == 'nan':
                continue

            self.tariffs[cat] = Tariff(
                category=cat,
                commission_rate=float(row.get('commission_rate', 0.15)),
                delivery_cost=float(row.get('delivery_cost', 120.0)),
                processing_cost=float(row.get('processing_cost', 45.0)),
                source="Загружено пользователем"
            )
            loaded += 1

        logger.info(f"Загружено {loaded} тарифов")
        return loaded

    def get_best_tariff(
        self,
        category_name: str,
        ym_api: Optional[YandexMarketAPI] = None,
        use_api: bool = True
    ) -> Tariff:
        """Получение оптимального тарифа с кэшированием."""
        cat_clean = category_name.lower().strip()
        cache_key = f"{cat_clean}_FBS"

        # Проверка LRU-кэша
        cached = self._cache.get(cache_key)
        if cached:
            return cached

        # Приоритет 1: API Яндекс Маркета
        if use_api and ym_api and ym_api.api_key:
            try:
                result = ym_api.calculate_tariffs(
                    [{
                        "categoryId": 0,
                        "price": 1000,
                        "length": 10,
                        "width": 10,
                        "height": 10,
                        "weight": 1,
                        "quantity": 1
                    }],
                    selling_program="FBS"
                )
                if result and len(result) > 0:
                    t = self._parse_ym_tariffs(
                        result[0].get("tariffs", []),
                        cat_clean
                    )
                    if t:
                        self._cache.set(cache_key, t)
                        return t
            except Exception as e:
                logger.warning(f"API ЯМ сбой для {cat_clean}: {e}")

        # Приоритет 2: Загруженный справочник
        if cat_clean in self.tariffs:
            t = self.tariffs[cat_clean]
            self._cache.set(cache_key, t)
            return t

        # Фоллбэк
        logger.warning(
            f"Тариф для '{cat_clean}' не найден. Применён фоллбэк 15%."
        )
        t = Tariff(
            category=cat_clean,
            commission_rate=0.15,
            source="⚠️ БАЗОВЫЙ ФОЛЛБЭК"
        )
        self._cache.set(cache_key, t)
        return t

    def _parse_ym_tariffs(
        self,
        tariffs: List[Dict],
        category: str
    ) -> Optional[Tariff]:
        """Парсинг тарифов из ответа API."""
        if not tariffs:
            return None

        commission_rate = 0.15
        delivery_cost = 120.0
        processing_cost = 45.0

        for t in tariffs:
            t_type = str(t.get('type', '')).upper()
            amount = float(t.get('amount', 0))
            if t_type == 'FEE':
                commission_rate = amount / 100.0 if amount > 1 else amount
            elif t_type == 'DELIVERY_TO_CUSTOMER':
                delivery_cost = amount
            elif t_type == 'SORTING':
                processing_cost = amount

        return Tariff(
            category=category,
            commission_rate=commission_rate,
            delivery_cost=delivery_cost,
            processing_cost=processing_cost,
            source="API Яндекс Маркета"
        )

    def get_tariffs_vectorized(
        self,
        df: pd.DataFrame,
        ym_api: Optional[YandexMarketAPI] = None,
        use_api: bool = True
    ) -> pd.DataFrame:
        """Векторизованное получение тарифов."""
        if hasattr(df['category'], 'cat'):
            unique_cats = df['category'].cat.categories
        else:
            unique_cats = df['category'].unique()

        tariff_map = {}
        for cat in unique_cats:
            tariff_map[cat] = self.get_best_tariff(cat, ym_api, use_api)

        tariff_df = pd.DataFrame([
            {'category': cat, **t.to_dict()}
            for cat, t in tariff_map.items()
        ])
        return tariff_df

# ============================================================================
# ВАЛИДАТОР ДАННЫХ (БЛОК 4)
# ============================================================================
class DataValidator:
    """Валидация и очистка входных данных."""

    REQUIRED_COLS: Final[Tuple[str, ...]] = (
        'artikul', 'category', 'selling_price', 'cogs'
    )
    NUMERIC_COLS: Final[Tuple[str, ...]] = (
        'selling_price', 'cogs', 'packaging_cost', 'delivery_cost',
        'processing_cost', 'other_costs', 'return_rate', 'return_cost_amount'
    )

    @classmethod
    def validate(cls, df: pd.DataFrame) -> Tuple[pd.DataFrame, List[str]]:
        """Валидация DataFrame."""
        errors = []

        if df.empty:
            return df, ["DataFrame пустой"]

        df_validated = df.copy()

        # Проверка обязательных колонок
        missing = [c for c in cls.REQUIRED_COLS if c not in df_validated.columns]
        if missing:
            errors.append(f"Отсутствуют обязательные колонки: {missing}")

        # Проверка числовых колонок
        for col in cls.NUMERIC_COLS:
            if col in df_validated.columns:
                df_validated[col] = pd.to_numeric(
                    df_validated[col],
                    errors='coerce'
                ).fillna(0).clip(lower=0)

        # Проверка selling_price
        if 'selling_price' in df_validated.columns:
            zero_prices = (df_validated['selling_price'] == 0).sum()
            if zero_prices > 0:
                errors.append(
                    f"selling_price: {zero_prices} SKU с нулевой ценой"
                )

        return df_validated, errors

# ============================================================================
# ФИНАНСОВЫЙ ДВИЖОК — СТРОГИЕ ФОРМУЛЫ FBS (БЛОК 5)
# ============================================================================
class FinancialEngine:
    """
    Векторизованный расчёт unit-экономики по строгим формулам FBS.

    Переменные:
        P       = selling_price       (Цена продажи)
        C       = cogs                (Себестоимость)
        Pack    = packaging_cost      (Упаковка)
        Com%    = commission_rate     (Комиссия Маркета по категории)
        Log     = delivery_cost       (Логистика FBS, ₽)
        Proc    = processing_cost     (Обработка отправления, ₽)
        Pay%    = acquiring_rate      (Эквайринг / приём платежа)
        Ad%     = advertising_rate    (Реклама / ДРР)
        Tax%    = tax_rate            (Налоговая ставка, УСН 6% от выручки)
        Ret%    = return_rate         (Доля возвратов / невыкупов)
        RetCost = return_cost_amount  (Средняя стоимость одного возврата)
        Other   = other_costs         (Прочие расходы)

    Формулы:
        Комиссия        = P × Com%
        Эквайринг       = P × Pay%
        Реклама         = P × Ad%
        Налог           = P × Tax%
        Риск возвратов  = Ret% × RetCost

        Прибыль = P × (1 − Com% − Pay% − Tax% − Ad%)
                  − C − Pack − Log − Proc − Other − Ret% × RetCost

        Маржинальность  = Прибыль / P × 100%
        Наценка         = Прибыль / C × 100%

        Pmin           = Fixed / (1 − Com% − Pay% − Tax% − Ad%)
        Цена (M=20%)   = Fixed / (1 − Com% − Pay% − Tax% − Ad% − 0.20)
        Цена (M=30%)   = Fixed / (1 − Com% − Pay% − Tax% − Ad% − 0.30)

        где Fixed = C + Pack + Log + Proc + Other + Ret% × RetCost
    """

    TARGET_MARGIN_20: Final[float] = 0.20
    TARGET_MARGIN_30: Final[float] = 0.30
    MIN_DENOMINATOR: Final[float] = 0.01

    @classmethod
    def calculate_all(
        cls,
        df: pd.DataFrame,
        tax_rate: float,
        acquiring_rate: float,
        advertising_rate: float,
        return_rate: float,
        return_cost: float,
        tariffs_map: Dict[str, Dict]
    ) -> pd.DataFrame:
        """Полный расчёт unit-экономики по формулам FBS."""
        # Копия для безопасности
        df = df.copy()

        # Исправление кодировки
        for col in ['artikul', 'category']:
            if col in df.columns:
                df[col] = df[col].astype(str).apply(StringUtils.fix_double_utf8)

        # Заполнение значений по умолчанию
        defaults = {
            'selling_price': 0.0,
            'cogs': 0.0,
            'packaging_cost': 0.0,
            'delivery_cost': 0.0,
            'processing_cost': 0.0,
            'other_costs': 0.0,
            'return_rate': return_rate,
            'return_cost_amount': return_cost
        }
        for col, default in defaults.items():
            if col not in df.columns:
                df[col] = default
            df[col] = pd.to_numeric(df[col], errors='coerce').fillna(default)

        # Если Ret% / RetCost не заданы построчно (нули) — берём глобальные
        df['return_rate'] = np.where(
            df['return_rate'] > 0, df['return_rate'], return_rate
        )
        df['return_cost_amount'] = np.where(
            df['return_cost_amount'] > 0, df['return_cost_amount'], return_cost
        )

        # Слияние с тарифами категорий (Com%, Log, Proc)
        tariff_df = pd.DataFrame.from_dict(tariffs_map, orient='index').reset_index()
        tariff_df.columns = ['category'] + list(tariff_df.columns[1:])

        df = df.merge(
            tariff_df[['category', 'commission_rate', 'delivery_cost', 'processing_cost']],
            on='category',
            how='left',
            suffixes=('', '_tariff')
        )

        # Логистика и обработка: приоритет — построчные значения из файла,
        # если они нулевые — берём из тарифа категории
        if 'delivery_cost_tariff' in df.columns:
            df['delivery_cost'] = np.where(
                df['delivery_cost'] > 0,
                df['delivery_cost'],
                df['delivery_cost_tariff'].fillna(120.0)
            )
            df = df.drop(columns=['delivery_cost_tariff'])

        if 'processing_cost_tariff' in df.columns:
            df['processing_cost'] = np.where(
                df['processing_cost'] > 0,
                df['processing_cost'],
                df['processing_cost_tariff'].fillna(45.0)
            )
            df = df.drop(columns=['processing_cost_tariff'])

        # Комиссия категории — всегда из тарифа
        df['commission_rate'] = pd.to_numeric(
            df['commission_rate'], errors='coerce'
        ).fillna(0.15)

        # ── Извлечение переменных как numpy-массивов ──────────────────────
        P = df['selling_price'].values.astype(np.float64)
        C = df['cogs'].values.astype(np.float64)
        Pack = df['packaging_cost'].values.astype(np.float64)
        Log = df['delivery_cost'].values.astype(np.float64)
        Proc = df['processing_cost'].values.astype(np.float64)
        Other = df['other_costs'].values.astype(np.float64)

        Com_pct = df['commission_rate'].values.astype(np.float64)
        Pay_pct = np.full_like(P, acquiring_rate)
        Tax_pct = np.full_like(P, tax_rate)
        Ad_pct = np.full_like(P, advertising_rate)

        Ret_pct = df['return_rate'].values.astype(np.float64)
        RetCost = df['return_cost_amount'].values.astype(np.float64)

        # ── 1. Процентные затраты от цены ─────────────────────────────────
        # Комиссия = P × Com%
        df['commission'] = P * Com_pct

        # Эквайринг = P × Pay%
        df['acquiring_cost'] = P * Pay_pct

        # Реклама = P × Ad%
        df['advertising_cost'] = P * Ad_pct

        # Налог = P × Tax% (УСН «Доходы»)
        df['tax_cost'] = P * Tax_pct

        # ── 2. Риск возвратов ─────────────────────────────────────────────
        # Ожидаемый расход на возвраты = Ret% × RetCost
        df['expected_return_cost'] = Ret_pct * RetCost

        # ── 3. Итоговые расходы ───────────────────────────────────────────
        df['total_expenses'] = (
            C + Pack + Log + Proc + Other +
            df['commission'].values +
            df['acquiring_cost'].values +
            df['advertising_cost'].values +
            df['tax_cost'].values +
            df['expected_return_cost'].values
        )

        # ── 4. Чистая прибыль (главная формула) ───────────────────────────
        # Прибыль = P × (1 − Com% − Pay% − Tax% − Ad%)
        #           − C − Pack − Log − Proc − Other − Ret% × RetCost
        multiplier = 1.0 - Com_pct - Pay_pct - Tax_pct - Ad_pct
        df['gross_profit'] = (
            P * multiplier
            - C - Pack - Log - Proc - Other
            - df['expected_return_cost'].values
        )

        # ── 5. Маржинальность и наценка ───────────────────────────────────
        # Маржинальность = Прибыль / P × 100%
        df['margin_percent'] = NumericUtils.safe_divide(
            df['gross_profit'].values, P, 0.0
        ) * 100.0

        # Наценка = Прибыль / C × 100%
        df['markup_percent'] = NumericUtils.safe_divide(
            df['gross_profit'].values, C, 0.0
        ) * 100.0

        # ── 6. Рекомендованные цены ───────────────────────────────────────
        # Fixed = C + Pack + Log + Proc + Other + Ret% × RetCost
        fixed_costs = C + Pack + Log + Proc + Other + df['expected_return_cost'].values

        # Pmin = Fixed / (1 − Com% − Pay% − Tax% − Ad%)
        denom_min = multiplier
        df['rec_price_min'] = np.where(
            denom_min > cls.MIN_DENOMINATOR,
            NumericUtils.safe_divide(fixed_costs, denom_min, np.nan),
            np.nan
        )

        # Цена для маржи 20% = Fixed / (1 − Com% − Pay% − Tax% − Ad% − 0.20)
        denom_20 = multiplier - cls.TARGET_MARGIN_20
        df['rec_price_20'] = np.where(
            denom_20 > cls.MIN_DENOMINATOR,
            NumericUtils.safe_divide(fixed_costs, denom_20, np.nan),
            np.nan
        )

        # Цена для маржи 30% = Fixed / (1 − Com% − Pay% − Tax% − Ad% − 0.30)
        denom_30 = multiplier - cls.TARGET_MARGIN_30
        df['rec_price_30'] = np.where(
            denom_30 > cls.MIN_DENOMINATOR,
            NumericUtils.safe_divide(fixed_costs, denom_30, np.nan),
            np.nan
        )

        # ── 7. Статус маржинальности ──────────────────────────────────────
        df['profitability_status'] = np.where(
            df['gross_profit'] > 0,
            np.where(
                df['margin_percent'] >= 20,
                'Высокомаржинальный',
                'Низкомаржинальный'
            ),
            'Убыточный'
        )

        # ── 8. Округление ─────────────────────────────────────────────────
        money_cols = [
            'commission', 'acquiring_cost', 'advertising_cost', 'tax_cost',
            'expected_return_cost', 'total_expenses', 'gross_profit',
            'rec_price_min', 'rec_price_20', 'rec_price_30'
        ]
        for col in money_cols:
            if col in df.columns:
                df[col] = NumericUtils.money_round(df[col].values)

        pct_cols = ['margin_percent', 'markup_percent']
        for col in pct_cols:
            if col in df.columns:
                df[col] = NumericUtils.percent_round(df[col].values)

        return df

# ============================================================================
# КЭШИРОВАННЫЙ РАСЧЁТ
# ============================================================================
@st.cache_data(ttl=CACHE_TTL, show_spinner=False)
def run_calculations_cached(
    df_hash: str,
    df: pd.DataFrame,
    tax_rate: float,
    acquiring_rate: float,
    advertising_rate: float,
    return_rate: float,
    return_cost: float,
    tariffs_map: Dict[str, Dict]
) -> pd.DataFrame:
    """Кэшированный расчёт с защитой от мутации."""
    if df.empty:
        return df

    result = FinancialEngine.calculate_all(
        df=df,
        tax_rate=tax_rate,
        acquiring_rate=acquiring_rate,
        advertising_rate=advertising_rate,
        return_rate=return_rate,
        return_cost=return_cost,
        tariffs_map=tariffs_map
    )

    # Оптимизация типов
    result = DtypeOptimizer.optimize(result)

    return result

# ============================================================================
# ЭКСПОРТ EXCEL — ULTRA DESIGN (БЛОК 6)
# ============================================================================
class UltimateExcelExporter:
    """
    Профессиональный Excel-экспорт с формулами FBS:
    - 📊 Дашборд: KPI-карточки, сводка по статусам, круговая диаграмма
    - 📋 Детальный расчёт: группировка колонок, статусные цвета, DataBar/ColorScale
    - 💡 Рекомендации: Pmin, цена для маржи 20% и 30%
    - 🧮 Формулы: справочный лист с формулами расчёта
    - ⚙️ Параметры: настройки расчёта
    """

    # ── Палитра ──────────────────────────────────────────────────────────
    C = {
        "navy":         "1F4E79",
        "blue":         "2E75B6",
        "light_blue":   "BDD7EE",
        "sky":          "DEEAF1",
        "green":        "375623",
        "lime":         "E2EFDA",
        "yellow":       "FFF2CC",
        "orange":       "FCE4D6",
        "crimson":      "C00000",
        "white":        "FFFFFF",
        "gray":         "F2F2F2",
        "dark_gray":    "595959",
        "profit_hi":    "375623",
        "profit_lo":    "7F6000",
        "loss":         "C00000",
        "profit_hi_bg": "E2EFDA",
        "profit_lo_bg": "FFEB9C",
        "loss_bg":      "FFC7CE",
    }

    # ── Числовые форматы ─────────────────────────────────────────────────
    FMT_MONEY: Final[str] = '#,##0.00 ₽'
    FMT_PCT: Final[str] = '0.00%'
    FMT_PCT1: Final[str] = '0.0%'
    FMT_INT: Final[str] = '#,##0'
    FMT_NUM2: Final[str] = '0.00'

    # ── Конструкторы стилей ──────────────────────────────────────────────
    @classmethod
    def _fill(cls, hex_color: str) -> PatternFill:
        return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

    @classmethod
    def _font(
        cls,
        bold: bool = False,
        color: str = "000000",
        size: int = 10,
        italic: bool = False,
        name: str = "Calibri"
    ) -> Font:
        return Font(bold=bold, color=color, size=size, italic=italic, name=name)

    @classmethod
    def _align(
        cls,
        h: str = "left",
        v: str = "center",
        wrap: bool = False
    ) -> Alignment:
        return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

    @classmethod
    def _thin_border(cls) -> Border:
        s = Side(style="thin", color="D9D9D9")
        return Border(left=s, right=s, top=s, bottom=s)

    @classmethod
    def _thick_border_bottom(cls) -> Border:
        s_thin = Side(style="thin", color="D9D9D9")
        s_med = Side(style="medium", color=cls.C["navy"])
        return Border(left=s_thin, right=s_thin, top=s_thin, bottom=s_med)

    # ── Главный метод экспорта ───────────────────────────────────────────
    @classmethod
    def export_max_info(
        cls,
        df: pd.DataFrame,
        tax_rate: float,
        acquiring_rate: float,
        advertising_rate: float,
        return_rate: float,
        return_cost: float
    ) -> bytes:
        """Экспорт с полным форматированием и формулами FBS."""
        if not OPENPYXL_AVAILABLE or df.empty:
            return b""

        # ── Подготовка данных ────────────────────────────────────────────
        df = df.copy()
        for col in df.select_dtypes(include=['category']).columns:
            df[col] = df[col].astype(str)

        wb = Workbook()

        STATUS_BG_MAP = {
            "Высокомаржинальный": cls.C["profit_hi_bg"],
            "Низкомаржинальный": cls.C["profit_lo_bg"],
            "Убыточный": cls.C["loss_bg"],
        }
        STATUS_FG_MAP = {
            "Высокомаржинальный": cls.C["profit_hi"],
            "Низкомаржинальный": cls.C["profit_lo"],
            "Убыточный": cls.C["loss"],
        }

        # ════════════════════════════════════════════════════════════════
        # ① ДАШБОРД
        # ════════════════════════════════════════════════════════════════
        ws_dash = wb.active
        ws_dash.title = "📊 Дашборд"
        ws_dash.sheet_view.showGridLines = False
        ws_dash.column_dimensions["A"].width = 2

        # -- Заголовок -----------------------------------------------------
        ws_dash.row_dimensions[1].height = 8
        ws_dash.row_dimensions[2].height = 40
        ws_dash.row_dimensions[3].height = 14

        ws_dash.merge_cells("B2:M2")
        title_cell = ws_dash["B2"]
        title_cell.value = "📊 UNIT-ECONOMICS FBS · ЯНДЕКС МАРКЕТ"
        title_cell.font = Font(name="Calibri", bold=True, size=20, color=cls.C["white"])
        title_cell.fill = cls._fill(cls.C["navy"])
        title_cell.alignment = cls._align(h="center")

        ws_dash.merge_cells("B3:M3")
        sub_cell = ws_dash["B3"]
        sub_cell.value = (
            f"Налог: {tax_rate * 100:.1f}%  ·  "
            f"Эквайринг: {acquiring_rate * 100:.1f}%  ·  "
            f"Реклама: {advertising_rate * 100:.1f}%  ·  "
            f"Возвраты: {return_rate * 100:.1f}% × {return_cost:.0f} ₽  ·  "
            f"Сформировано: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
        )
        sub_cell.font = Font(name="Calibri", italic=True, size=10, color=cls.C["white"])
        sub_cell.fill = cls._fill(cls.C["blue"])
        sub_cell.alignment = cls._align(h="center")

        # -- KPI-карточки (строки 5-9) ---------------------------------------
        ws_dash.row_dimensions[4].height = 10

        total_sku = len(df)
        avg_margin = float(df["margin_percent"].mean()) if "margin_percent" in df.columns else 0.0
        avg_markup = float(df["markup_percent"].mean()) if "markup_percent" in df.columns else 0.0
        profitable_cnt = int((df["gross_profit"] > 0).sum()) if "gross_profit" in df.columns else 0
        loss_cnt = int((df["gross_profit"] < 0).sum()) if "gross_profit" in df.columns else 0
        total_profit = float(df["gross_profit"].sum()) if "gross_profit" in df.columns else 0.0

        kpi_cards = [
            ("Всего SKU", total_sku, cls.FMT_INT, cls.C["navy"], cls.C["sky"], "🏷️"),
            ("Средняя маржа", avg_margin / 100.0, cls.FMT_PCT1, cls.C["green"], cls.C["lime"], "📈"),
            ("Прибыльных", profitable_cnt, cls.FMT_INT, cls.C["green"], cls.C["lime"], "✅"),
            ("Убыточных", loss_cnt, cls.FMT_INT, cls.C["crimson"], cls.C["orange"], "❌"),
            ("Средняя наценка", avg_markup / 100.0, cls.FMT_PCT1, cls.C["blue"], cls.C["light_blue"], "💰"),
            ("Сумма прибыли/ед.", total_profit, cls.FMT_MONEY, cls.C["navy"], cls.C["sky"], "💵"),
        ]

        card_cols = [
            ("B", "C"),
            ("D", "E"),
            ("F", "G"),
            ("H", "I"),
            ("J", "K"),
            ("L", "M"),
        ]

        for (label, value, fmt, fg, bg, icon), (start_col, end_col) in zip(kpi_cards, card_cols):
            sc = ord(start_col) - ord("A") + 1
            ec = ord(end_col) - ord("A") + 1

            for r in range(5, 10):
                for c in range(sc, ec + 1):
                    cell = ws_dash.cell(row=r, column=c)
                    cell.fill = cls._fill(bg)
                    cell.border = cls._thin_border()

            ws_dash.merge_cells(start_row=5, start_column=sc, end_row=5, end_column=ec)
            lbl = ws_dash.cell(row=5, column=sc, value=f"{icon} {label}")
            lbl.font = Font(name="Calibri", bold=True, size=9, color=fg)
            lbl.fill = cls._fill(bg)
            lbl.alignment = cls._align(h="center")

            ws_dash.merge_cells(start_row=6, start_column=sc, end_row=8, end_column=ec)
            val_cell = ws_dash.cell(row=6, column=sc, value=value)
            val_cell.font = Font(name="Calibri", bold=True, size=20, color=fg)
            val_cell.fill = cls._fill(bg)
            val_cell.number_format = fmt
            val_cell.alignment = cls._align(h="center", v="center")

        ws_dash.row_dimensions[5].height = 18
        ws_dash.row_dimensions[6].height = 34
        ws_dash.row_dimensions[7].height = 34
        ws_dash.row_dimensions[8].height = 18

        # -- Сводная таблица по статусам -------------------------------------
        ws_dash.row_dimensions[10].height = 10

        status_header_row = 11
        ws_dash.merge_cells(f"B{status_header_row}:F{status_header_row}")
        h = ws_dash.cell(row=status_header_row, column=2, value="Структура портфеля по маржинальности")
        h.font = cls._font(bold=True, color="FFFFFF", size=11)
        h.fill = cls._fill(cls.C["navy"])
        h.alignment = cls._align(h="center")

        if "profitability_status" in df.columns and "gross_profit" in df.columns:
            status_summary = (
                df.groupby("profitability_status", observed=True)
                .agg(
                    SKU=("artikul", "count"),
                    Прибыль=("gross_profit", "sum"),
                    Маржа_avg=("margin_percent", "mean"),
                )
                .reset_index()
                .rename(columns={"profitability_status": "Статус"})
            )

            col_hdrs = ["Статус", "SKU, шт.", "Сум. прибыль, ₽", "Ср. маржа, %"]

            for ci, hdr in enumerate(col_hdrs, 2):
                cell = ws_dash.cell(row=status_header_row + 1, column=ci, value=hdr)
                cell.font = cls._font(bold=True, color="FFFFFF", size=9)
                cell.fill = cls._fill(cls.C["blue"])
                cell.alignment = cls._align(h="center")
                cell.border = cls._thin_border()

            for ri, row_data in enumerate(status_summary.itertuples(), 2):
                fg = STATUS_FG_MAP.get(row_data.Статус, "000000")
                bg = STATUS_BG_MAP.get(row_data.Статус, "FFFFFF")
                row_idx = status_header_row + ri

                values = [
                    row_data.Статус,
                    int(row_data.SKU),
                    float(row_data.Прибыль),
                    float(row_data.Маржа_avg) / 100.0 if row_data.Маржа_avg else 0.0,
                ]
                fmts = [None, cls.FMT_INT, cls.FMT_MONEY, cls.FMT_PCT1]

                for ci, (val, fmt2) in enumerate(zip(values, fmts), 2):
                    cell = ws_dash.cell(row=row_idx, column=ci, value=val)
                    cell.fill = cls._fill(bg)
                    cell.font = Font(name="Calibri", bold=(ci == 2), color=fg, size=10)
                    cell.alignment = cls._align(h="center" if ci > 2 else "left")
                    cell.border = cls._thin_border()
                    if fmt2:
                        cell.number_format = fmt2

        # -- Круговая диаграмма по статусам ----------------------------------
        if "profitability_status" in df.columns:
            try:
                pc = PieChart()
                pc.title = "Структура по статусам"
                pc.width = 14
                pc.height = 10

                status_counts = df["profitability_status"].value_counts()
                hidden_start = 20
                ws_dash.cell(row=hidden_start, column=17, value="Статус")
                ws_dash.cell(row=hidden_start, column=18, value="Кол-во")
                for i, (st_name, cnt) in enumerate(status_counts.items(), 1):
                    ws_dash.cell(row=hidden_start + i, column=17, value=st_name)
                    ws_dash.cell(row=hidden_start + i, column=18, value=int(cnt))

                ws_dash.column_dimensions["Q"].hidden = True
                ws_dash.column_dimensions["R"].hidden = True

                data_ref = Reference(
                    ws_dash,
                    min_col=18,
                    min_row=hidden_start,
                    max_row=hidden_start + len(status_counts)
                )
                labels_ref = Reference(
                    ws_dash,
                    min_col=17,
                    min_row=hidden_start + 1,
                    max_row=hidden_start + len(status_counts)
                )
                pc.add_data(data_ref, titles_from_data=True)
                pc.set_categories(labels_ref)

                dl = DataLabelList()
                dl.showPercent = True
                dl.showVal = False
                pc.dataLabels = dl

                ws_dash.add_chart(pc, "H11")
            except Exception as chart_err:
                logger.warning(f"PieChart: {chart_err}")

        # ════════════════════════════════════════════════════════════════
        # ② ДЕТАЛЬНЫЙ РАСЧЁТ
        # ════════════════════════════════════════════════════════════════
        ws_det = wb.create_sheet("📋 Детальный расчёт")
        ws_det.sheet_view.showGridLines = False
        ws_det.freeze_panes = "C3"

        # Группировка колонок по блокам тарифов FBS
        COLUMN_GROUPS: List[Tuple[str, List[str], str]] = [
            ("🏷️ Товар", ["artikul", "category"], cls.C["navy"]),
            ("💰 Цены", ["selling_price", "cogs"], cls.C["blue"]),
            ("📦 Прямые затраты", ["packaging_cost", "delivery_cost", "processing_cost", "other_costs"], "375623"),
            ("💳 Процентные удержания", ["commission", "acquiring_cost", "advertising_cost", "tax_cost"], "366092"),
            ("🔄 Возвраты", ["return_rate", "return_cost_amount", "expected_return_cost"], "7F6000"),
            ("📈 Финрезультат", ["total_expenses", "gross_profit", "margin_percent", "markup_percent"], cls.C["green"]),
            ("🎯 Рекомендации", ["rec_price_min", "rec_price_20", "rec_price_30"], "833C00"),
            ("📊 Статус", ["profitability_status"], "1F4E79"),
        ]

        COL_DISPLAY_NAMES: Dict[str, str] = {
            "artikul": "Артикул",
            "category": "Категория",
            "selling_price": "Цена продажи (P)",
            "cogs": "Себестоимость (C)",
            "packaging_cost": "Упаковка (Pack)",
            "delivery_cost": "Логистика FBS (Log)",
            "processing_cost": "Обработка (Proc)",
            "other_costs": "Прочие (Other)",
            "commission": "Комиссия = P×Com%",
            "acquiring_cost": "Эквайринг = P×Pay%",
            "advertising_cost": "Реклама = P×Ad%",
            "tax_cost": "Налог = P×Tax%",
            "return_rate": "Доля возвратов (Ret%)",
            "return_cost_amount": "Стоимость возврата (RetCost)",
            "expected_return_cost": "Риск возвратов = Ret%×RetCost",
            "total_expenses": "Итого расходы",
            "gross_profit": "Чистая прибыль",
            "margin_percent": "Маржа, %",
            "markup_percent": "Наценка, %",
            "rec_price_min": "Pmin (цена в ноль)",
            "rec_price_20": "Цена для маржи 20%",
            "rec_price_30": "Цена для маржи 30%",
            "profitability_status": "Статус",
        }

        COL_FORMATS: Dict[str, str] = {
            "selling_price": cls.FMT_MONEY,
            "cogs": cls.FMT_MONEY,
            "packaging_cost": cls.FMT_MONEY,
            "delivery_cost": cls.FMT_MONEY,
            "processing_cost": cls.FMT_MONEY,
            "other_costs": cls.FMT_MONEY,
            "commission": cls.FMT_MONEY,
            "acquiring_cost": cls.FMT_MONEY,
            "advertising_cost": cls.FMT_MONEY,
            "tax_cost": cls.FMT_MONEY,
            "return_rate": cls.FMT_PCT,
            "return_cost_amount": cls.FMT_MONEY,
            "expected_return_cost": cls.FMT_MONEY,
            "total_expenses": cls.FMT_MONEY,
            "gross_profit": cls.FMT_MONEY,
            "margin_percent": cls.FMT_NUM2,
            "markup_percent": cls.FMT_NUM2,
            "rec_price_min": cls.FMT_MONEY,
            "rec_price_20": cls.FMT_MONEY,
            "rec_price_30": cls.FMT_MONEY,
        }

        # Упорядоченный список колонок из групп
        ordered_cols: List[str] = []
        for _, cols, _ in COLUMN_GROUPS:
            for c in cols:
                if c in df.columns and c not in ordered_cols:
                    ordered_cols.append(c)

        df_export = df[ordered_cols].copy()

        # ─ Строка 1: группы (цветные мерджи) ────────────────────────────
        group_row = 1
        current_col = 1
        for group_name, group_cols, group_color in COLUMN_GROUPS:
            valid_cols = [c for c in group_cols if c in df_export.columns]
            if not valid_cols:
                continue
            span = len(valid_cols)
            end_col = current_col + span - 1

            if span > 1:
                ws_det.merge_cells(
                    start_row=group_row, start_column=current_col,
                    end_row=group_row, end_column=end_col
                )
            cell = ws_det.cell(row=group_row, column=current_col, value=group_name)
            cell.fill = cls._fill(group_color)
            cell.font = cls._font(bold=True, color="FFFFFF", size=9)
            cell.alignment = cls._align(h="center")
            cell.border = cls._thick_border_bottom()

            current_col = end_col + 1

        ws_det.row_dimensions[group_row].height = 22

        # ─ Строка 2: названия колонок ────────────────────────────────────
        header_row = 2
        for ci, col in enumerate(df_export.columns, 1):
            cell = ws_det.cell(
                row=header_row, column=ci,
                value=COL_DISPLAY_NAMES.get(col, col)
            )
            cell.fill = cls._fill(cls.C["navy"])
            cell.font = cls._font(bold=True, color="FFFFFF", size=9)
            cell.alignment = cls._align(h="center", wrap=True)
            cell.border = cls._thin_border()

        ws_det.row_dimensions[header_row].height = 34

        # ─ Данные ────────────────────────────────────────────────────────
        data_start_row = 3
        n_rows = len(df_export)
        n_cols = len(df_export.columns)
        col_name_list = list(df_export.columns)

        for ri, (_, row_data) in enumerate(df_export.iterrows()):
            r_idx = data_start_row + ri
            zebra = cls.C["gray"] if ri % 2 == 0 else cls.C["white"]
            status = str(row_data.get("profitability_status", ""))
            row_bg = STATUS_BG_MAP.get(status, zebra)

            for ci, col in enumerate(col_name_list, 1):
                val = row_data[col]
                if isinstance(val, (np.floating, np.integer)):
                    val = float(val)
                cell = ws_det.cell(row=r_idx, column=ci, value=val)
                cell.border = cls._thin_border()
                cell.alignment = cls._align(
                    h="right" if col in COL_FORMATS else "left",
                    v="center"
                )

                if col in COL_FORMATS:
                    cell.fill = cls._fill(row_bg)
                    cell.number_format = COL_FORMATS[col]
                    if col == "gross_profit" and isinstance(val, (int, float)) and val < 0:
                        cell.font = cls._font(bold=True, color=cls.C["crimson"])
                    else:
                        cell.font = cls._font(size=10)
                elif col == "profitability_status":
                    cell.fill = cls._fill(STATUS_BG_MAP.get(status, cls.C["white"]))
                    cell.font = Font(
                        name="Calibri", bold=True, size=9,
                        color=STATUS_FG_MAP.get(status, "000000")
                    )
                    cell.alignment = cls._align(h="center")
                else:
                    cell.fill = cls._fill(zebra)
                    cell.font = cls._font(size=10)

        # ─ Условное форматирование: DataBar для маржи ────────────────────
        if "margin_percent" in col_name_list:
            m_idx = col_name_list.index("margin_percent") + 1
            m_col_letter = get_column_letter(m_idx)
            m_range = f"{m_col_letter}{data_start_row}:{m_col_letter}{data_start_row + n_rows - 1}"
            try:
                db_rule = DataBarRule(
                    start_type="num", start_value=-50,
                    end_type="num", end_value=50,
                    color=cls.C["blue"]
                )
                ws_det.conditional_formatting.add(m_range, db_rule)
            except Exception as e:
                logger.warning(f"DataBar rule: {e}")

        # ─ Условное форматирование: ColorScale для прибыли ────────────────
        if "gross_profit" in col_name_list:
            p_idx = col_name_list.index("gross_profit") + 1
            p_col_letter = get_column_letter(p_idx)
            p_range = f"{p_col_letter}{data_start_row}:{p_col_letter}{data_start_row + n_rows - 1}"
            try:
                cs_rule = ColorScaleRule(
                    start_type="min", start_color="FFC7CE",
                    mid_type="num", mid_value=0, mid_color="FFEB9C",
                    end_type="max", end_color="C6EFCE",
                )
                ws_det.conditional_formatting.add(p_range, cs_rule)
            except Exception as e:
                logger.warning(f"ColorScale rule: {e}")

        # ─ Автоширина ────────────────────────────────────────────────────
        MIN_WIDTH, MAX_WIDTH = 10, 32
        for ci, col in enumerate(col_name_list, 1):
            col_letter = get_column_letter(ci)
            sample_vals = df_export[col].astype(str).str.len()
            best_len = max(
                len(COL_DISPLAY_NAMES.get(col, col)) * 0.75,
                float(sample_vals.quantile(0.95)) if not sample_vals.empty else 0.0
            )
            ws_det.column_dimensions[col_letter].width = min(
                max(best_len + 2, MIN_WIDTH), MAX_WIDTH
            )

        # ─ Excel-таблица с автофильтрами ──────────────────────────────────
        table_ref = f"A{header_row}:{get_column_letter(n_cols)}{data_start_row + n_rows - 1}"
        try:
            tab = Table(displayName="UnitEconomicsFBS", ref=table_ref)
            tab.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium2",
                showRowStripes=True,
                showFirstColumn=False,
                showLastColumn=False,
            )
            ws_det.add_table(tab)
        except Exception as te:
            logger.warning(f"Table add error: {te}")

        # ════════════════════════════════════════════════════════════════
        # ③ ЛИСТ РЕКОМЕНДАЦИЙ
        # ════════════════════════════════════════════════════════════════
        ws_rec = wb.create_sheet("💡 Рекомендации")
        ws_rec.sheet_view.showGridLines = False

        ws_rec.merge_cells("A1:H1")
        rec_title = ws_rec["A1"]
        rec_title.value = "💡 ЦЕНОВЫЕ РЕКОМЕНДАЦИИ (Pmin / маржа 20% / маржа 30%)"
        rec_title.font = cls._font(bold=True, color="FFFFFF", size=14)
        rec_title.fill = cls._fill(cls.C["navy"])
        rec_title.alignment = cls._align(h="center")
        ws_rec.row_dimensions[1].height = 32

        rec_cols = [
            "artikul", "category", "selling_price", "cogs",
            "gross_profit", "margin_percent",
            "rec_price_min", "rec_price_20", "rec_price_30",
            "profitability_status"
        ]
        rec_cols = [c for c in rec_cols if c in df.columns]
        df_rec = df[rec_cols].copy()

        REC_HEADER_NAMES = {
            "artikul": "Артикул",
            "category": "Категория",
            "selling_price": "Текущая цена (P)",
            "cogs": "Себестоимость (C)",
            "gross_profit": "Прибыль / ед.",
            "margin_percent": "Маржа, %",
            "rec_price_min": "▶ Pmin (в ноль)",
            "rec_price_20": "▶ Цена маржа 20%",
            "rec_price_30": "▶ Цена маржа 30%",
            "profitability_status": "Статус",
        }
        REC_FMTS = {
            "selling_price": cls.FMT_MONEY,
            "cogs": cls.FMT_MONEY,
            "gross_profit": cls.FMT_MONEY,
            "margin_percent": cls.FMT_NUM2,
            "rec_price_min": cls.FMT_MONEY,
            "rec_price_20": cls.FMT_MONEY,
            "rec_price_30": cls.FMT_MONEY,
        }

        hdr_row = 2
        for ci, col in enumerate(rec_cols, 1):
            cell = ws_rec.cell(row=hdr_row, column=ci, value=REC_HEADER_NAMES.get(col, col))
            cell.fill = cls._fill(cls.C["blue"])
            cell.font = cls._font(bold=True, color="FFFFFF", size=9)
            cell.alignment = cls._align(h="center", wrap=True)
            cell.border = cls._thin_border()
        ws_rec.row_dimensions[hdr_row].height = 30

        for ri, (_, row_data) in enumerate(df_rec.iterrows()):
            r_idx = hdr_row + 1 + ri
            status = str(row_data.get("profitability_status", ""))
            zebra = cls.C["gray"] if ri % 2 == 0 else cls.C["white"]

            for ci, col in enumerate(rec_cols, 1):
                val = row_data[col]
                if isinstance(val, (np.floating, np.integer)):
                    val = float(val)
                cell = ws_rec.cell(row=r_idx, column=ci, value=val)
                cell.border = cls._thin_border()

                if col in REC_FMTS:
                    cell.number_format = REC_FMTS[col]
                    cell.alignment = cls._align(h="right")
                elif col == "profitability_status":
                    cell.alignment = cls._align(h="center")
                else:
                    cell.alignment = cls._align(h="left")

                if col == "profitability_status":
                    cell.fill = cls._fill(STATUS_BG_MAP.get(status, cls.C["white"]))
                    cell.font = Font(
                        name="Calibri", bold=True, size=9,
                        color=STATUS_FG_MAP.get(status, "000000")
                    )
                elif col in ("rec_price_min", "rec_price_20", "rec_price_30"):
                    cell.fill = cls._fill(cls.C["sky"])
                    cell.font = cls._font(bold=True, color=cls.C["navy"])
                else:
                    cell.fill = cls._fill(zebra)
                    cell.font = cls._font(size=10)

        for ci, col in enumerate(rec_cols, 1):
            cl = get_column_letter(ci)
            header_len = len(REC_HEADER_NAMES.get(col, col)) + 2
            if not df_rec.empty:
                data_len = float(df_rec[col].astype(str).str.len().quantile(0.9)) + 2
            else:
                data_len = 10.0
            ws_rec.column_dimensions[cl].width = min(max(header_len, data_len), 28)

        ws_rec.freeze_panes = "C3"

        # ════════════════════════════════════════════════════════════════
        # ④ ЛИСТ ФОРМУЛ (СПРАВОЧНИК)
        # ════════════════════════════════════════════════════════════════
        ws_frm = wb.create_sheet("🧮 Формулы")
        ws_frm.sheet_view.showGridLines = False

        ws_frm.merge_cells("A1:C1")
        frm_title = ws_frm["A1"]
        frm_title.value = "🧮 ФОРМУЛЫ РАСЧЁТА ЮНИТ-ЭКОНОМИКИ FBS"
        frm_title.font = cls._font(bold=True, color="FFFFFF", size=14)
        frm_title.fill = cls._fill(cls.C["navy"])
        frm_title.alignment = cls._align(h="center")
        ws_frm.row_dimensions[1].height = 32

        formulas_list = [
            ("Показатель", "Формула", "Комментарий"),
            ("Комиссия Маркета", "P × Com%", "Процент от цены, зависит от категории товара"),
            ("Эквайринг", "P × Pay%", "Приём и перевод платежей; 0, если включён в тарифы"),
            ("Реклама (ДРР)", "P × Ad%", "Буст продаж, продвижение как % от цены"),
            ("Налог (УСН Доходы)", "P × Tax%", "Например, 6% от выручки"),
            ("Риск возвратов", "Ret% × RetCost", "Ожидаемый расход на возвраты на каждую продажу"),
            ("Итого расходы", "C + Pack + Log + Proc + Other + P×(Com%+Pay%+Ad%+Tax%) + Ret%×RetCost", "Сумма всех затрат на единицу"),
            ("Чистая прибыль", "P × (1 − Com% − Pay% − Tax% − Ad%) − C − Pack − Log − Proc − Other − Ret% × RetCost", "Главная формула прибыли на единицу"),
            ("Маржинальность", "Прибыль / P × 100%", "Доля прибыли в цене продажи"),
            ("Наценка", "Прибыль / C × 100%", "Отношение прибыли к себестоимости"),
            ("Pmin (цена в ноль)", "(C + Pack + Log + Proc + Other + Ret%×RetCost) / (1 − Com% − Pay% − Tax% − Ad%)", "Ниже этой цены продавать нельзя"),
            ("Цена для маржи 20%", "(C + Pack + Log + Proc + Other + Ret%×RetCost) / (1 − Com% − Pay% − Tax% − Ad% − 0.20)", "Цена для целевой маржинальности 20%"),
            ("Цена для маржи 30%", "(C + Pack + Log + Proc + Other + Ret%×RetCost) / (1 − Com% − Pay% − Tax% − Ad% − 0.30)", "Цена для целевой маржинальности 30%"),
        ]

        for ri, (name, formula, comment) in enumerate(formulas_list, 2):
            is_header = (ri == 2)
            bg = cls.C["blue"] if is_header else (cls.C["sky"] if ri % 2 == 0 else cls.C["white"])
            fg = "FFFFFF" if is_header else "000000"

            c1 = ws_frm.cell(row=ri, column=1, value=name)
            c1.font = cls._font(bold=True, color=fg, size=10)
            c1.fill = cls._fill(bg)
            c1.alignment = cls._align(h="left", wrap=True)
            c1.border = cls._thin_border()

            c2 = ws_frm.cell(row=ri, column=2, value=formula)
            c2.font = Font(name="Consolas", bold=is_header, color=fg, size=10)
            c2.fill = cls._fill(bg)
            c2.alignment = cls._align(h="left", wrap=True)
            c2.border = cls._thin_border()

            c3 = ws_frm.cell(row=ri, column=3, value=comment)
            c3.font = cls._font(bold=is_header, color=fg, size=9, italic=not is_header)
            c3.fill = cls._fill(bg)
            c3.alignment = cls._align(h="left", wrap=True)
            c3.border = cls._thin_border()

            ws_frm.row_dimensions[ri].height = 30

        ws_frm.column_dimensions["A"].width = 24
        ws_frm.column_dimensions["B"].width = 75
        ws_frm.column_dimensions["C"].width = 48

        # ════════════════════════════════════════════════════════════════
        # ⑤ ПАРАМЕТРЫ РАСЧЁТА
        # ════════════════════════════════════════════════════════════════
        ws_par = wb.create_sheet("⚙️ Параметры")
        ws_par.sheet_view.showGridLines = False

        ws_par.merge_cells("A1:B1")
        p_title = ws_par["A1"]
        p_title.value = "⚙️ ПАРАМЕТРЫ РАСЧЁТА"
        p_title.font = cls._font(bold=True, color="FFFFFF", size=13)
        p_title.fill = cls._fill(cls.C["navy"])
        p_title.alignment = cls._align(h="center")
        ws_par.row_dimensions[1].height = 28

        params_list = [
            ("Версия приложения", APP_VERSION),
            ("Дата формирования", datetime.now().strftime("%d.%m.%Y %H:%M:%S")),
            ("Модель работы", "FBS (склад продавца)"),
            ("Налог (Tax%)", f"{tax_rate * 100:.2f}%"),
            ("Эквайринг (Pay%)", f"{acquiring_rate * 100:.2f}%"),
            ("Реклама / ДРР (Ad%)", f"{advertising_rate * 100:.2f}%"),
            ("Доля возвратов (Ret%)", f"{return_rate * 100:.2f}%"),
            ("Стоимость возврата (RetCost)", f"{return_cost:.2f} ₽"),
            ("Всего SKU", len(df)),
            ("Прибыльных SKU", int((df["gross_profit"] > 0).sum()) if "gross_profit" in df.columns else "—"),
            ("Убыточных SKU", int((df["gross_profit"] < 0).sum()) if "gross_profit" in df.columns else "—"),
            ("Средняя маржа", f"{float(df['margin_percent'].mean()):.2f}%" if "margin_percent" in df.columns else "—"),
            ("Средняя наценка", f"{float(df['markup_percent'].mean()):.2f}%" if "markup_percent" in df.columns else "—"),
        ]

        for ri, (param, val) in enumerate(params_list, 2):
            bg = cls.C["sky"] if ri % 2 == 0 else cls.C["white"]

            p_cell = ws_par.cell(row=ri, column=1, value=param)
            p_cell.font = cls._font(bold=True, size=10)
            p_cell.fill = cls._fill(bg)
            p_cell.alignment = cls._align(h="left")
            p_cell.border = cls._thin_border()

            v_cell = ws_par.cell(row=ri, column=2, value=val)
            v_cell.font = cls._font(size=10)
            v_cell.fill = cls._fill(bg)
            v_cell.alignment = cls._align(h="left")
            v_cell.border = cls._thin_border()

        ws_par.column_dimensions["A"].width = 32
        ws_par.column_dimensions["B"].width = 40

        # ── Сохранение ────────────────────────────────────────────────────
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

# ============================================================================
# НОРМАЛИЗАТОР ДАННЫХ (БЛОК 7)
# ============================================================================
class UniversalDataNormalizer:
    """Нормализация входных данных под переменные формул FBS."""

    COLUMN_MAPPING: Final[Dict[str, List[str]]] = {
        'artikul': ['artikul', 'артикул', 'sku', 'offer_id', 'id', 'код'],
        'category': ['category', 'категория', 'группа', 'предмет', 'тип'],
        'selling_price': ['selling_price', 'цена продажи', 'цена', 'price', 'стоимость', 'p'],
        'cogs': ['cogs', 'себестоимость', 'закупка', 'cost', 'закупочная', 'c'],
        'packaging_cost': ['packaging_cost', 'упаковка', 'стоимость упаковки', 'pack'],
        'delivery_cost': ['delivery_cost', 'логистика', 'доставка', 'логистика fbs', 'log'],
        'processing_cost': ['processing_cost', 'обработка', 'сортировка', 'обработка отправления', 'proc'],
        'other_costs': ['other_costs', 'прочие', 'прочие расходы', 'маркировка', 'other'],
        'return_rate': ['return_rate', 'доля возвратов', 'возвраты %', 'ret%'],
        'return_cost_amount': ['return_cost_amount', 'стоимость возврата', 'цена возврата', 'retcost'],
    }

    NUMERIC_COLS: Final[Tuple[str, ...]] = (
        'selling_price', 'cogs', 'packaging_cost', 'delivery_cost',
        'processing_cost', 'other_costs', 'return_rate', 'return_cost_amount'
    )

    @classmethod
    def normalize_dataframe(cls, raw_df: pd.DataFrame) -> pd.DataFrame:
        """Нормализация DataFrame."""
        if raw_df.empty:
            return raw_df

        df = raw_df.copy()
        df.columns = [str(col).strip().lower() for col in df.columns]

        # Поиск колонок по синонимам
        final_data = {}
        for target_col, synonyms in cls.COLUMN_MAPPING.items():
            found = False
            for syn in synonyms:
                if syn in df.columns:
                    final_data[target_col] = df[syn]
                    found = True
                    break

            if not found:
                if target_col == 'artikul':
                    final_data[target_col] = [
                        f"SKU_{i+1}" for i in range(len(df))
                    ]
                elif target_col == 'category':
                    final_data[target_col] = "не указано"
                else:
                    final_data[target_col] = 0.0

        norm_df = pd.DataFrame(final_data)

        # Очистка числовых колонок
        for col in cls.NUMERIC_COLS:
            if col in norm_df.columns:
                norm_df[col] = pd.to_numeric(
                    norm_df[col].astype(str).str.replace(
                        r'[\s,;%₽]', '', regex=True
                    ),
                    errors='coerce'
                ).fillna(0.0).abs()

        # Если return_rate указан в процентах (например, 5 вместо 0.05) — нормализуем
        if 'return_rate' in norm_df.columns:
            norm_df['return_rate'] = np.where(
                norm_df['return_rate'] > 1.0,
                norm_df['return_rate'] / 100.0,
                norm_df['return_rate']
            )

        # Очистка строковых колонок
        norm_df['artikul'] = norm_df['artikul'].astype(str).str.strip()
        norm_df['category'] = norm_df['category'].astype(str).str.strip().str.lower()

        # Удаление дубликатов
        return norm_df.drop_duplicates(subset=['artikul'], keep='first')

    @classmethod
    def load_file(
        cls,
        file_buffer: io.BytesIO,
        file_name: str
    ) -> pd.DataFrame:
        """Загрузка файла с автоопределением формата."""
        try:
            if file_name.endswith('.csv'):
                return pd.read_csv(
                    file_buffer,
                    sep=None,
                    engine='python',
                    encoding='utf-8'
                )
            elif file_name.endswith(('.xls', '.xlsx')):
                return pd.read_excel(file_buffer)
            else:
                raise ValueError("Неподдерживаемый формат файла")
        except UnicodeDecodeError:
            file_buffer.seek(0)
            return pd.read_csv(
                file_buffer,
                sep=None,
                engine='python',
                encoding='cp1251'
            )

# ============================================================================
# ОРКЕСТРАТОР ПАЙПЛАЙНА (БЛОК 8)
# ============================================================================
@dataclass
class DataPipeline:
    """Оркестратор пайплайна обработки данных."""

    tax_rate: float = 0.06
    acquiring_rate: float = 0.01
    advertising_rate: float = 0.05
    return_rate: float = 0.05
    return_cost: float = 150.0
    use_api: bool = False

    def process(
        self,
        raw_df: pd.DataFrame,
        tariff_manager: HybridTariffManager,
        ym_api: Optional[YandexMarketAPI] = None
    ) -> pd.DataFrame:
        """Полный пайплайн обработки."""
        # Нормализация
        norm_df = UniversalDataNormalizer.normalize_dataframe(raw_df)

        # Валидация
        validated_df, errors = DataValidator.validate(norm_df)

        if errors:
            logger.warning(f"Ошибки валидации: {errors}")

        # Получение тарифов категорий (Com%, Log, Proc)
        tariff_df = tariff_manager.get_tariffs_vectorized(
            validated_df,
            ym_api=ym_api,
            use_api=self.use_api
        )
        tariffs_map = tariff_df.set_index('category').to_dict(orient='index')

        # Расчёт
        current_hash = StringUtils.make_hash(validated_df)

        calc_df = run_calculations_cached(
            df_hash=current_hash,
            df=validated_df,
            tax_rate=self.tax_rate,
            acquiring_rate=self.acquiring_rate,
            advertising_rate=self.advertising_rate,
            return_rate=self.return_rate,
            return_cost=self.return_cost,
            tariffs_map=tariffs_map
        )

        return calc_df

# ============================================================================
# STREAMLIT UI (БЛОК 9)
# ============================================================================
def init_session_state():
    """Инициализация состояния сессии."""
    defaults = {
        'main_df': pd.DataFrame(),
        'calc_df': pd.DataFrame(),
        'tariffs': {},
        'last_hash': '',
        'api_key': '',
        'business_id': ''
    }
    for key, val in defaults.items():
        if key not in st.session_state:
            st.session_state[key] = val


def render_sidebar() -> DataPipeline:
    """Рендеринг сайдбара с настройками ставок для формул."""
    with st.sidebar:
        st.header("⚙️ Ставки для формул FBS")

        st.markdown("**Процентные удержания от цены (P):**")

        tax_rate = st.number_input(
            "Налог Tax% (УСН Доходы)",
            min_value=0.0,
            max_value=30.0,
            value=6.0,
            step=0.5,
            help="Налог = P × Tax%. Для УСН «Доходы» обычно 6%."
        ) / 100.0

        acquiring_rate = st.number_input(
            "Эквайринг Pay%",
            min_value=0.0,
            max_value=10.0,
            value=1.0,
            step=0.1,
            help="Эквайринг = P × Pay%. Если включён в тарифы — поставьте 0."
        ) / 100.0

        advertising_rate = st.number_input(
            "Реклама / ДРР Ad%",
            min_value=0.0,
            max_value=50.0,
            value=5.0,
            step=0.5,
            help="Реклама = P × Ad%. Доля рекламных расходов от цены."
        ) / 100.0

        st.markdown("---")
        st.markdown("**Возвраты и невыкупы:**")

        return_rate = st.number_input(
            "Доля возвратов Ret%",
            min_value=0.0,
            max_value=50.0,
            value=5.0,
            step=0.5,
            help="Ожидаемый расход на возвраты = Ret% × RetCost."
        ) / 100.0

        return_cost = st.number_input(
            "Стоимость одного возврата RetCost, ₽",
            min_value=0.0,
            max_value=5000.0,
            value=150.0,
            step=10.0,
            help="Обратная логистика + обработка возврата."
        )

        st.markdown("---")
        st.markdown("**API Яндекс Маркета (опционально):**")

        api_key = st.text_input(
            "API Key Яндекс Маркета",
            value=st.session_state.api_key,
            type="password"
        )
        business_id = st.text_input(
            "Business ID",
            value=st.session_state.business_id
        )
        st.session_state.api_key = api_key
        st.session_state.business_id = business_id

        use_api = st.checkbox(
            "Использовать API ЯМ для тарифов категорий",
            value=False
        )

        st.markdown("---")
        st.caption(
            "Главная формула:  \n"
            "**Прибыль = P × (1 − Com% − Pay% − Tax% − Ad%) − C − Pack − Log − Proc − Other − Ret% × RetCost**"
        )

    return DataPipeline(
        tax_rate=tax_rate,
        acquiring_rate=acquiring_rate,
        advertising_rate=advertising_rate,
        return_rate=return_rate,
        return_cost=return_cost,
        use_api=use_api
    )


def render_upload_section() -> Tuple[Optional[pd.DataFrame], HybridTariffManager]:
    """Рендеринг секции загрузки данных."""
    col1, col2 = st.columns(2)
    tariff_manager = HybridTariffManager()
    main_df = None

    with col1:
        st.subheader("1. Загрузка данных товаров")
        st.caption(
            "Колонки: артикул, категория, цена (P), себестоимость (C), "
            "упаковка (Pack), логистика (Log), обработка (Proc), прочие (Other). "
            "Незаполненные Log/Proc берутся из тарифа категории."
        )
        uploaded_file = st.file_uploader(
            "Загрузите Excel или CSV",
            type=['xlsx', 'xls', 'csv']
        )

        if uploaded_file is not None:
            try:
                raw_df = UniversalDataNormalizer.load_file(
                    io.BytesIO(uploaded_file.read()),
                    uploaded_file.name
                )
                norm_df = UniversalDataNormalizer.normalize_dataframe(raw_df)
                validated_df, errors = DataValidator.validate(norm_df)

                if errors:
                    for err in errors:
                        st.warning(err)

                main_df = validated_df
                st.session_state.main_df = validated_df
                st.success(
                    f"Загружено {len(validated_df)} уникальных SKU"
                )
            except Exception as e:
                st.error(f"Ошибка чтения файла: {e}")

    with col2:
        st.subheader("2. Справочник тарифов категорий (опционально)")
        st.caption(
            "Колонки: category, commission_rate (Com%), "
            "delivery_cost (Log, ₽), processing_cost (Proc, ₽). "
            "Актуальные ставки — в кабинете продавца: Финансы → Тарифы → FBS."
        )
        tariff_file = st.file_uploader(
            "Тарифы (Excel/CSV)",
            type=['xlsx', 'xls', 'csv'],
            key="tariff_uploader"
        )

        if tariff_file is not None:
            try:
                t_raw = UniversalDataNormalizer.load_file(
                    io.BytesIO(tariff_file.read()),
                    tariff_file.name
                )
                t_raw.columns = [str(c).strip().lower() for c in t_raw.columns]
                tariff_manager.load_tariffs_from_file(t_raw)
                st.success(
                    f"Загружено {len(tariff_manager.tariffs)} тарифов"
                )
            except Exception as e:
                st.error(f"Ошибка тарифов: {e}")

    return main_df, tariff_manager


def render_results(df_calc: pd.DataFrame) -> None:
    """Рендеринг результатов расчёта."""
    st.subheader("3. Результаты расчёта")

    # Сводные метрики
    c1, c2, c3, c4, c5 = st.columns(5)
    c1.metric("Всего SKU", len(df_calc))
    c2.metric(
        "Средняя маржа, %",
        f"{df_calc['margin_percent'].mean():.1f}%"
    )
    c3.metric(
        "Средняя наценка, %",
        f"{df_calc['markup_percent'].mean():.1f}%"
    )
    c4.metric(
        "Прибыльных SKU",
        len(df_calc[df_calc['gross_profit'] > 0])
    )
    c5.metric(
        "Убыточных SKU",
        len(df_calc[df_calc['gross_profit'] < 0])
    )

    # Фильтр
    status_filter = st.multiselect(
        "Фильтр по статусу",
        options=list(df_calc['profitability_status'].unique()),
        default=list(df_calc['profitability_status'].unique())
    )

    filtered_df = df_calc[
        df_calc['profitability_status'].isin(status_filter)
    ]

    # Таблица с ключевыми показателями по формулам
    display_cols = [
        'artikul', 'category', 'selling_price', 'cogs',
        'commission', 'acquiring_cost', 'advertising_cost', 'tax_cost',
        'expected_return_cost', 'total_expenses',
        'gross_profit', 'margin_percent', 'markup_percent',
        'rec_price_min', 'rec_price_20', 'rec_price_30',
        'profitability_status'
    ]
    display_cols = [c for c in display_cols if c in filtered_df.columns]

    st.dataframe(
        filtered_df[display_cols],
        use_container_width=True,
        height=420,
        column_config={
            "artikul": st.column_config.TextColumn("Артикул"),
            "category": st.column_config.TextColumn("Категория"),
            "selling_price": st.column_config.NumberColumn("Цена (P)", format="%.2f ₽"),
            "cogs": st.column_config.NumberColumn("Себестоимость (C)", format="%.2f ₽"),
            "commission": st.column_config.NumberColumn("Комиссия P×Com%", format="%.2f ₽"),
            "acquiring_cost": st.column_config.NumberColumn("Эквайринг P×Pay%", format="%.2f ₽"),
            "advertising_cost": st.column_config.NumberColumn("Реклама P×Ad%", format="%.2f ₽"),
            "tax_cost": st.column_config.NumberColumn("Налог P×Tax%", format="%.2f ₽"),
            "expected_return_cost": st.column_config.NumberColumn("Возвраты Ret%×RetCost", format="%.2f ₽"),
            "total_expenses": st.column_config.NumberColumn("Итого расходы", format="%.2f ₽"),
            "gross_profit": st.column_config.NumberColumn("Чистая прибыль", format="%.2f ₽"),
            "margin_percent": st.column_config.NumberColumn("Маржа, %", format="%.2f"),
            "markup_percent": st.column_config.NumberColumn("Наценка, %", format="%.2f"),
            "rec_price_min": st.column_config.NumberColumn("Pmin (в ноль)", format="%.2f ₽"),
            "rec_price_20": st.column_config.NumberColumn("Цена маржа 20%", format="%.2f ₽"),
            "rec_price_30": st.column_config.NumberColumn("Цена маржа 30%", format="%.2f ₽"),
            "profitability_status": st.column_config.TextColumn("Статус"),
        }
    )


def render_export_section(
    filtered_df: pd.DataFrame,
    pipeline: DataPipeline
) -> None:
    """Рендеринг секции экспорта."""
    st.subheader("4. Экспорт")

    excel_data = UltimateExcelExporter.export_max_info(
        df=filtered_df,
        tax_rate=pipeline.tax_rate,
        acquiring_rate=pipeline.acquiring_rate,
        advertising_rate=pipeline.advertising_rate,
        return_rate=pipeline.return_rate,
        return_cost=pipeline.return_cost
    )

    if excel_data:
        st.download_button(
            label="📥 Скачать Excel (Дашборд · Расчёт · Рекомендации · Формулы · Параметры)",
            data=excel_data,
            file_name=f"fbs_unit_economics_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    else:
        st.warning(
            "Не удалось сформировать Excel. Проверьте установку openpyxl."
        )


def main():
    """Главная функция приложения."""
    st.set_page_config(
        page_title=APP_NAME,
        page_icon="📈",
        layout="wide"
    )
    init_session_state()

    st.title(f"📊 {APP_NAME} v{APP_VERSION}")
    st.markdown(
        "Калькулятор юнит-экономики **FBS** по строгим формулам Яндекс Маркета. "
        "Главная формула: **Прибыль = P × (1 − Com% − Pay% − Tax% − Ad%) − C − Pack − Log − Proc − Other − Ret% × RetCost**"
    )

    # Настройки
    pipeline = render_sidebar()

    # Загрузка данных
    main_df, tariff_manager = render_upload_section()

    # Расчёт
    st.markdown("---")
    if main_df is not None and not main_df.empty:
        if st.button("🚀 Рассчитать юнит-экономику FBS", type="primary"):
            with st.spinner("Выполняется векторизованный расчёт по формулам FBS..."):
                ym_api = YandexMarketAPI(
                    api_key=st.session_state.api_key,
                    business_id=st.session_state.business_id
                ) if st.session_state.api_key else None

                calc_df = pipeline.process(main_df, tariff_manager, ym_api)
                st.session_state.calc_df = calc_df
                st.session_state.last_hash = StringUtils.make_hash(main_df)

                st.success("Расчёт завершён успешно!")

    # Отображение результатов (сохраняется между перерисовками)
    if not st.session_state.calc_df.empty:
        render_results(st.session_state.calc_df)
        render_export_section(st.session_state.calc_df, pipeline)


if __name__ == "__main__":
    main()
```

## Что интегрировано и что удалено

### ✅ Интегрированные формулы (расчёты идентичны в приложении и экспорте)

| Показатель | Формула | Столбец |
|---|---|---|
| Комиссия | `P × Com%` | `commission` |
| Эквайринг | `P × Pay%` | `acquiring_cost` |
| Реклама (ДРР) | `P × Ad%` | `advertising_cost` |
| Налог (УСН 6%) | `P × Tax%` | `tax_cost` |
| Риск возвратов | `Ret% × RetCost` | `expected_return_cost` |
| Чистая прибыль | `P×(1−Com%−Pay%−Tax%−Ad%) − C − Pack − Log − Proc − Other − Ret%×RetCost` | `gross_profit` |
| Маржинальность | `Прибыль / P × 100%` | `margin_percent` |
| Наценка | `Прибыль / C × 100%` | `markup_percent` |
| Pmin (в ноль) | `Fixed / (1−Com%−Pay%−Tax%−Ad%)` | `rec_price_min` |
| Цена для маржи 20% | `Fixed / (1−Com%−Pay%−Tax%−Ad%−0.20)` | `rec_price_20` |
| Цена для маржи 30% | `Fixed / (1−Com%−Pay%−Tax%−Ad%−0.30)` | `rec_price_30` |

### ❌ Удалённые столбцы и формулы
- Объёмный/оплачиваемый вес, габариты (`weight_kg`, `length_cm`, `width_cm`, `height_cm`, `billable_weight`)
- Магистраль (`middle_mile_cost`), спецтариф (`is_special_tariff`)
- ABC/XYZ анализ, `daily_sales`, `stock_depth_days`, `warehouse_cost`
- Сложные тарифные сетки доставки (`delivery_min`, `delivery_max`, `delivery_rate`)
- Точка безубыточности в штуках, `contribution_margin`

### 📄 Листы в Excel-экспорте
1. **📊 Дашборд** — KPI-карточки, сводка по статусам, круговая диаграмма
2. **📋 Детальный расчёт** — группировка колонок по блокам тарифов ЯМ, статусные цвета, DataBar и ColorScale
3. **💡 Рекомендации** — Pmin, цены для маржи 20% и 30%
4. **🧮 Формулы** — справочный лист со всеми применёнными формулами
5. **⚙️ Параметры** — ставки Tax%, Pay%, Ad%, Ret%, RetCost и сводные метрики
